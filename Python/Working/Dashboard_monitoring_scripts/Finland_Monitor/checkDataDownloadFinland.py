# -*- coding: utf-8 -*-
"""
Created on Mon Jan  2 10:30:05 2023

@author: krishnan
"""

import os
import pandas as pd
import time
import openpyxl
from openpyxl import load_workbook
import xlwt
import xlrd
from xlutils.copy import copy

pathSpecified = "N:/COVerAGE-DB/Automation/Finland"  #Location from where the current day's automation data has to be extracted
timestr = time.strftime("%Y%m%d")


files=pd.read_csv('N:/COVerAGE-DB/Automation/Finland/FinlandRows.csv') #CSV File where the previous day's automation data is stored
print(len(files))
list_of_files=os.listdir(pathSpecified) # Extracting current day's automation data
print(len(list_of_files))

wbkName = 'N:/COVerAGE-DB/Automation/Python-Dashboard/PythonDashboard.xlsx' #Indicating the path and filename of Python Dashboard
wbk = openpyxl.load_workbook(wbkName) #Opening the Python Dashboard
wks = wbk["dash-view"] #Choosing the first worksheet in the excel workbook
wks.cell(row=3, column=6).value = wks.cell(row=3, column=5).value #Writing the previous day's record
#myRow = 9
for myCol in range(14,7,-1):
    wks.cell(row=3, column=myCol).value = wks.cell(row=3, column=myCol-1).value
#wks.cell(row=9, column=14).value = wks.cell(row=9, column=13).value
#wks.cell(row=9, column=13).value = wks.cell(row=9, column=12).value
#wks.cell(row=9, column=12).value = wks.cell(row=9, column=11).value
#wks.cell(row=9, column=11).value = wks.cell(row=9, column=10).value
#wks.cell(row=9, column=10).value = wks.cell(row=9, column=9).value
#wks.cell(row=9, column=9).value = wks.cell(row=9, column=8).value
#wks.cell(row=9, column=8).value = wks.cell(row=9, column=7).value
wks.cell(row=3, column=7).value = "FALSE"
wbk.save(wbkName)
wbk.close

if (len(files))!=len(list_of_files): #Compare previous day's data in csv file and current day's data
    
    list_of_files=list(set(list_of_files))
    pd.DataFrame({'files':list_of_files}).to_csv('N:/COVerAGE-DB/Automation/Finland/FinlandRows.csv') #Updating the Previous day's CSV Files with current day's records
    print(len(list_of_files))
    print(len(files))
    wbkName = 'N:/COVerAGE-DB/Automation/Python-Dashboard/PythonDashboard.xlsx' #Indicating the path and filename of Python Dashboard
    wbk = openpyxl.load_workbook(wbkName) #Opening the Python Dashboard
    wks = wbk["dash-view"] #Choosing the first worksheet in the excel workbook
    wks.cell(row=3, column=5).value = len(list_of_files) #Writing the current day's record in Dashboard
    wks.cell(row=3, column=7).value = "TRUE" #Indicate DATA DOWNLOADED ON THE PARTICULAR DATE
    wbk.save(wbkName) # Saving and closing the Dashboard
    wbk.close